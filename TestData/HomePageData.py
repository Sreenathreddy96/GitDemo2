import openpyxl


class HomePageData:
    test_HomePage_Data = [
        {"firstname": "Rahul", "E-mail": "hello@gmail.com", "password": "123456",
         "gender": "Male"},
        {"firstname": "Anshika", "E-mail": "hello@gmail.com", "password": "123456", "gender": "Female"}]

    @staticmethod
    def foo():
        #Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Sreenath\\Documents\\TesCases.xlsx")
        sheet = book.active
        final_list = []
        for row in range(2, sheet.max_row + 1):
            Name = sheet['A' + str(row)].value
            firstname = sheet['B' + str(row)].value
            Email = sheet['C' + str(row)].value
            password = sheet['D' + str(row)].value
            gender = sheet['E' + str(row)].value

            final_list.append({'Name': Name,'firstname': firstname,'Email': Email, 'password': password,'gender': gender })
        return final_list
        #for i in range(1, sheet.max_row + 1):

            #if sheet.cell(row=i, column=1).value == test_case_name:
                #for j in range(2, sheet.max_column + 1):
                    #Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return final_list
